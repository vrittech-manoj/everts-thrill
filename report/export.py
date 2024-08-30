# from django.shortcuts import render
# from rest_framework.response import Response
# from rest_framework.views import APIView
# from rest_framework import status
# from django.http import HttpResponse
# import csv
# from django.apps import apps
# from drf_yasg.utils import swagger_auto_schema
# from drf_yasg import openapi

# class GetSampleAPIView(APIView):
#     """
#     API view to generate and download sample data in CSV format for different models.
#     """
    
#     model_mapping = {
#         "package": "destination.Package",
#         "destination": "destination.Destination",
#         "gallery-images": "destination.DestinationGalleryImages",
#         "review": "review.Review",
#         "collection": "collection.Collection",
#         "departure": "departure.Departure",
#         "destination-book": "booking.DestinationBook",
#         "activity": "activities.Activity",
#     }

#     excluded_fields = ['id', 'slug', 'public_id', 'created_date', 'updated_date']

#     @swagger_auto_schema(
#         operation_description="Get a sample CSV for the specified model type.",
#         manual_parameters=[
#             openapi.Parameter(
#                 'type',
#                 openapi.IN_PATH,
#                 description="Specify the type of data to fetch as CSV (e.g., 'package', 'destination').",
#                 type=openapi.TYPE_STRING,
#                 enum=list(model_mapping.keys())
#             )
#         ],
#         responses={200: 'CSV file', 400: 'Unknown type'}
#     )
#     def get(self, request, type, format=None):
#         model_class_path = self.model_mapping.get(type)

#         if not model_class_path:
#             return Response({"message": 'Unknown type'}, status=status.HTTP_400_BAD_REQUEST)

#         # Dynamically import the model class
#         app_label, model_name = model_class_path.split(".")
#         model_class = apps.get_model(app_label=app_label, model_name=model_name)

#         # Get all fields of the model, excluding specified fields
#         column_list = []
#         for field in model_class._meta.get_fields():
#             if field.name not in self.excluded_fields:
#                 if field.is_relation:
#                     if field.many_to_one or field.one_to_one:
#                         # For ForeignKey or OneToOneField, include the related model's primary key or string representation
#                         column_list.append(f"{field.name}_names")
#                     elif field.many_to_many:
#                         # For ManyToManyField, append related objects' names
#                         column_list.append(f"{field.name}_names")
#                 else:
#                     column_list.append(field.name)

#         queryset = model_class.objects.all()

#         response = HttpResponse(content_type='text/csv')
#         response['Content-Disposition'] = f'attachment; filename="{type}.csv"'

#         writer = csv.writer(response)

#         # Write the header row
#         writer.writerow(column_list)

#         # Write data rows
#         for data in queryset:
#             data_row = []
#             for column in column_list:
#                 if column.endswith("_id"):
#                     # ForeignKey or OneToOneField relation, get related object's ID
#                     related_field_name = column.replace("_id", "")
#                     related_object = getattr(data, related_field_name)
#                     data_row.append(getattr(related_object, 'id', None) if related_object else None)
#                 elif column.endswith("_names"):
#                     # ManyToManyField relation, get related objects' names
#                     related_field_name = column.replace("_names", "")
#                     related_objects = getattr(data, related_field_name).all()
#                     related_names = ", ".join(str(obj) for obj in related_objects)
#                     data_row.append(related_names)
#                 else:
#                     data_row.append(getattr(data, column))
#             writer.writerow(data_row)

#         return response


import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from django.apps import apps
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

class GetSampleAPIView(APIView):
    """
    API view to generate and download sample data in Excel format for different models,
    including dropdowns for selecting related model data.
    """
    
    model_mapping = {
        "package": "destination.Package",
        "destination": "destination.Destination",
        "gallery-images": "destination.DestinationGalleryImages",
        "review": "review.Review",
        "collection": "collection.Collection",
        "departure": "departure.Departure",
        "destination-book": "booking.DestinationBook",
        "activity": "activities.Activity",
    }

    excluded_fields = ['id', 'slug', 'public_id', 'created_date', 'updated_date']

    @swagger_auto_schema(
        operation_description="Get a sample Excel file for the specified model type.",
        manual_parameters=[
            openapi.Parameter(
                'type',
                openapi.IN_PATH,
                description="Specify the type of data to fetch as Excel (e.g., 'package', 'destination').",
                type=openapi.TYPE_STRING,
                enum=list(model_mapping.keys())
            )
        ],
        responses={200: 'Excel file', 400: 'Unknown type'}
    )
    def get(self, request, type, format=None):
        model_class_path = self.model_mapping.get(type)

        if not model_class_path:
            return Response({"message": 'Unknown type'}, status=status.HTTP_400_BAD_REQUEST)

        # Dynamically import the model class
        app_label, model_name = model_class_path.split(".")
        model_class = apps.get_model(app_label=app_label, model_name=model_name)

        # Get all fields of the model, excluding specified fields and fields containing "image"
        column_list = []
        related_models = {}
        for field in model_class._meta.get_fields():
            if field.name not in self.excluded_fields and "image" not in field.name:
                if field.is_relation:
                    related_model = field.related_model
                    related_models[field.name] = related_model
                    # For dropdowns, we use names instead of IDs
                    column_list.append(f"{field.name}")
                else:
                    column_list.append(field.name)

        # Create an Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = model_name

        # Write the header row
        for col_num, column_title in enumerate(column_list, 1):
            sheet.cell(row=1, column=col_num, value=column_title)

        # Fetch and write data, ensuring names are used instead of IDs for related fields
        for row_num, data in enumerate(model_class.objects.all(), 2):
            for col_num, column in enumerate(column_list, 1):
                if column in related_models:
                    related_objs = getattr(data, column)
                    if hasattr(related_objs, 'all'):  # ManyToManyField
                        value = ", ".join([str(obj) for obj in related_objs.all()])
                    else:  # ForeignKey or OneToOneField
                        value = str(related_objs) if related_objs else None
                else:
                    value = getattr(data, column)
                sheet.cell(row=row_num, column=col_num, value=value)

        # Add dropdowns for ForeignKey and ManyToMany fields using related models
        for col_num, (field_name, related_model) in enumerate(related_models.items(), 1):
            related_objs = related_model.objects.all()
            related_names = [str(obj) for obj in related_objs]

            # Add related names to a new sheet for data validation
            validation_sheet_name = f"{related_model._meta.model_name}_data"
            if validation_sheet_name not in workbook.sheetnames:
                validation_sheet = workbook.create_sheet(title=validation_sheet_name)
                for i, name in enumerate(related_names, 1):
                    validation_sheet.cell(row=i, column=1, value=name)

            dv = DataValidation(
                type="list",
                formula1=f'{validation_sheet_name}!$A$1:$A${len(related_names)}',
                showDropDown=True
            )
            dv.error = 'Invalid entry, please select from the list'
            dv.errorTitle = 'Invalid Entry'

            # Apply the dropdown to the correct column in the main sheet using get_column_letter
            col_letter = get_column_letter(column_list.index(field_name) + 1)
            dv_range = f"{col_letter}2:{col_letter}{len(model_class.objects.all()) + 1}"
            dv.add(dv_range)
            sheet.add_data_validation(dv)

        # Prepare the response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{type}.xlsx"'

        # Save the workbook to the response
        workbook.save(response)

        return response

